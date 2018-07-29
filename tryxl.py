import openpyxl
import sys
import os

# To open already existing excel file
work_book = openpyxl.load_workbook('example.xlsx')
# workbook.get_sheet_names() : gives names of all sheets
sheet = work_book.get_sheet_by_name('Sheet1')
cell1 = sheet['A1']
print(str(cell1.value))
# To print values in the excel sheet
for i in range(1, 8):
	print(sheet.cell(row=i, column=2).value)
# Creation of a new excel workbook
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.get_sheet_by_name('Sheet')
new_sheet['A1'] = "hello"
# creating a new worksheet
new_sheet2 = new_workbook.create_sheet()
new_sheet2['A1'] = "world"
# Saving new workbook in tryedit.xsxl file
new_workbook.save('tryedit.xlsx')
